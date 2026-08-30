import openpyxl, json, re
from openpyxl.utils import get_column_letter, column_index_from_string

SRC = '/home/oleg/Downloads/Сводная Таблица по ТМ 35 от 26.08.26.xlsx'
wb = openpyxl.load_workbook(SRC, data_only=True)

# Tabs that carry a раздел/конструкция × вид-работ status matrix (folder-status
# trackers). "Скользячки" (dimensions register: T1/T2/diameter) and
# "Подкладки на cк. опоры (ОПн)" (schemes register) do NOT fit this pattern —
# verified by inspection (survey_out.json), excluded here, not guessed.
TABS = [
    ('opn', 'ОПН', 'ОПН'),
    ('opn_rsm', 'ОПН (рсм)', 'ОПН (рсм)'),
    ('opv', 'ОПВ', 'ОПВ'),
    ('n', 'Н', 'Н'),
    ('izolyaciya', 'Изоляция', 'Изоляция'),
    ('truba_svarka', 'Труба Сварка', 'Труба Сварка'),
    ('lotki', 'Лотки', 'Лотки'),
    ('sodk', 'СОДК', 'СОДК'),
    ('skl_o', 'Скользящие опоры', 'Скл О'),
    ('no', 'НО в лотках', 'НО'),
    ('kamery', 'Камеры', 'Камеры'),
    ('transhei', 'Траншеи', 'Траншеи'),
    ('kolodcy_1_16', 'Колодцы 1-16', 'Колодцы 1-16'),
    ('kolodcy_17_25', 'Колодцы 17-25', 'Колодцы 17-25'),
    ('obvyazka', 'Обвязка', 'Обвязка'),
    ('elektrika', 'Электрика', 'Электрика'),
    ('pavilyony', 'Павильоны', 'Павильоны'),
    ('met_konstr', 'Мет констр', 'Мет констр'),
]

NUM_HDR_RE = re.compile(r'^\s*(\d+)\.\s*(.+?)\s*$')


def clean(v):
    if v is None:
        return None
    if isinstance(v, str):
        v = v.replace('\n', ' ').strip()
        v = re.sub(r'\s+', ' ', v)
        return v or None
    return v


def parse_tab(code, label, sheet_name):
    ws = wb[sheet_name]
    maxc = ws.max_column

    # locate header row containing numbered work-type headers ("1. ...");
    # fall back to the row with the most non-empty text cells beyond col D
    # (unnumbered single/few work-type tabs: СОДК, Скл О, Камеры, Колодцы...)
    header_row = None
    for r in (1, 2):
        for c in range(1, maxc + 1):
            v = clean(ws.cell(row=r, column=c).value)
            if v and NUM_HDR_RE.match(v):
                header_row = r
                break
        if header_row:
            break
    numbered_headers = bool(header_row)
    if not header_row:
        # Prefer row1 (macro work-stage names, e.g. "Снятие почвенно-
        # растительного слоя" or a single "СОДК"/"Монтаж трубопровода")
        # over row2 (document-type sub-labels like "АОСР"/"ИС"/"Паспорта"
        # repeated per stage) — row2 only used if row1 is truly empty.
        for r in (1, 2):
            cnt = sum(1 for c in range(4, maxc + 1) if clean(ws.cell(row=r, column=c).value))
            if cnt:
                header_row = r
                break
    if not header_row:
        return {'code': code, 'error': 'no header row found at all'}

    meta_row = header_row + 1   # "Подписант"/"АОСР" etc
    resp_row = header_row + 3   # "Ответственный" (skip a numeric-index row in between, as on ОПН)
    # Some tabs put "Ответственный" straight after meta_row (no numeric-index row) —
    # detect by scanning rows header_row+1..header_row+4 for a row whose first
    # non-empty cell (col A/B/C) contains "ответств"
    resp_row = None
    for r in range(header_row + 1, min(header_row + 5, ws.max_row + 1)):
        for c in range(1, min(5, maxc + 1)):
            v = clean(ws.cell(row=r, column=c).value)
            if v and isinstance(v, str) and 'ответств' in v.lower():
                resp_row = r
                break
        if resp_row:
            break

    # identity columns: scan header_row and the rows immediately around it
    # (some tabs put "Раздел/Наименование конструкции" one row ABOVE the
    # numbered work-type headers — ОПН; others put it on the SAME row —
    # ОПН (рсм) and most others)
    id_cols = {}
    for r in (header_row - 1, header_row, header_row + 1):
        if r < 1:
            continue
        for c in range(1, min(8, maxc + 1)):
            v = clean(ws.cell(row=r, column=c).value)
            if not v or not isinstance(v, str):
                continue
            low = v.lower()
            if '№ п/п' in low or low.startswith('№п/п') or low.startswith('№ п'):
                id_cols.setdefault('num', c)
            elif 'раздел' in low or ('участ' in low and 'учас' not in id_cols):
                id_cols.setdefault('section', c)
            elif 'наименование конструкции' in low:
                id_cols.setdefault('construction', c)
            elif 'наименование фундамента' in low or 'тип фундамента' in low:
                id_cols.setdefault('foundation', c)

    has_section_level = 'section' in id_cols
    id_col = id_cols.get('construction') or id_cols.get('section')

    # work-type blocks: numbered headers in header_row -> column ranges;
    # tabs without numbering (СОДК, Скл О, Камеры, Колодцы, Обвязка) get
    # each non-empty header_row cell beyond the identity columns treated
    # as its own single-column work type (no sub-columns to group).
    id_col_set = set(id_cols.values())
    numbered = []
    if numbered_headers:
        for c in range(1, maxc + 1):
            v = clean(ws.cell(row=header_row, column=c).value)
            if v:
                m = NUM_HDR_RE.match(v)
                if m:
                    numbered.append((c, m.group(1), m.group(2)))
    else:
        # meta/tracking columns, not construction stages — appear right
        # after identity columns on nearly every tab in this pattern
        SKIP_LABELS = {'передано на проверку в рск', 'реестр к папке'}
        n = 0
        for c in range(1, maxc + 1):
            if c in id_col_set:
                continue
            v = clean(ws.cell(row=header_row, column=c).value)
            if v and isinstance(v, str) and v.lower().strip() not in SKIP_LABELS:
                n += 1
                numbered.append((c, str(n), v))
    work_types = []
    for i, (c, num, name) in enumerate(numbered):
        next_c = numbered[i + 1][0] if i + 1 < len(numbered) else maxc + 1
        signer = clean(ws.cell(row=meta_row, column=c).value) if meta_row <= ws.max_row else None
        resp = clean(ws.cell(row=resp_row, column=c).value) if resp_row else None
        wt_label = name if not numbered_headers else f'{num}. {name}'
        work_types.append({
            'name': wt_label, 'order': int(num), 'col': get_column_letter(c),
            'col_end': next_c - 1, 'responsible': resp, 'signer': signer,
        })

    # data rows: start after resp_row (or meta_row if no resp_row), stop at
    # first row where id_col is empty for 3 consecutive rows or hits a
    # "ответств"/"статус" marker (reference block below table)
    data_start = (resp_row or meta_row) + 1
    rows = []
    r = data_start
    blank_streak = 0
    while r <= ws.max_row and blank_streak < 3:
        val = clean(ws.cell(row=r, column=id_col).value) if id_col else None
        if val and isinstance(val, str) and ('ответств' in val.lower() or 'статус' in val.lower()):
            break
        if val in (None, ''):
            blank_streak += 1
            r += 1
            continue
        blank_streak = 0
        section_val = clean(ws.cell(row=r, column=id_cols['section']).value) if 'section' in id_cols else None
        construction_val = clean(ws.cell(row=r, column=id_cols['construction']).value) if 'construction' in id_cols else val
        foundation_val = clean(ws.cell(row=r, column=id_cols['foundation']).value) if 'foundation' in id_cols else None
        rows.append({
            'section': section_val, 'construction': construction_val,
            'foundation': foundation_val, 'source_row': r,
        })
        r += 1

    # carry forward section label to rows where it's blank (merged-cell pattern)
    last_section = None
    for row in rows:
        if row['section']:
            last_section = row['section']
        elif has_section_level:
            row['section'] = last_section

    return {
        'code': code, 'label': label, 'sheet': sheet_name,
        'has_section_level': has_section_level,
        'header_row': header_row, 'meta_row': meta_row, 'resp_row': resp_row,
        'id_cols': {k: get_column_letter(v) for k, v in id_cols.items()},
        'work_types': work_types,
        'rows': rows,
        'row_count': len(rows),
        'unique_sections': len({r['section'] for r in rows if r['section']}),
    }


# Status reference — PER TAB, not global (found by direct inspection,
# 28.08.2026): 16 of 18 tabs share one 8-item list ending in the
# distinctive label "Нет лаборатории"; only "ОПН" uses a different,
# 9-item RSK-review-cycle list. Detected by anchoring on that label
# text rather than fixed row numbers (row numbers differ per tab).
COMMON_STATUS_ANCHOR = 'Нет лаборатории'


def find_status_block(sheet_name):
    ws = wb[sheet_name]
    if sheet_name == 'ОПН':
        statuses = []
        for r in range(103, 112):
            code = clean(ws.cell(row=r, column=2).value)
            label = clean(ws.cell(row=r, column=3).value)
            if code and label:
                statuses.append({'code': code, 'label': label, 'order': r - 102})
        return statuses, 'own-block(103-111)'

    anchor_row = None
    for r in range(1, ws.max_row + 1):
        for c in (1, 2):
            if clean(ws.cell(row=r, column=c).value) == COMMON_STATUS_ANCHOR:
                anchor_row = r
                break
        if anchor_row:
            break
    if not anchor_row:
        return None, 'not-found'

    start = anchor_row - 7
    statuses = []
    for i, r in enumerate(range(start, anchor_row + 1)):
        code = clean(ws.cell(row=r, column=1).value)
        label = clean(ws.cell(row=r, column=2).value)
        if label:
            statuses.append({'code': code or label, 'label': label, 'order': i + 1})
    return statuses, f'common-block({start}-{anchor_row})'


result = {'tabs': []}
for code, label, sheet in TABS:
    parsed = parse_tab(code, label, sheet)
    statuses, block_loc = find_status_block(sheet)
    parsed['statuses'] = statuses
    parsed['status_block_location'] = block_loc
    result['tabs'].append(parsed)

with open('/tmp/claude-1000/-home-oleg/d973c3f1-942e-40b1-9945-4f57a68a3420/scratchpad/id_excel_parsed.json', 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

for t in result['tabs']:
    if 'error' in t:
        print(f"{t['code']:16s} ERROR: {t['error']}")
    else:
        n_st = len(t['statuses']) if t['statuses'] else 0
        print(f"{t['code']:16s} work_types={len(t['work_types']):3d} rows={t['row_count']:4d} "
              f"sections={t['unique_sections']:3d} statuses={n_st} ({t['status_block_location']})")
