#!/usr/bin/env python3
"""
Batch-импорт графика работ ТМ-35 из Excel в нормализованные таблицы.

Структура файла и алгоритм разбора описаны в
.claude/skills/tm35-excel/SKILL.md — читать перед правкой этого файла.

Выход: JSON-файлы в import/output/ (works.json, daily_progress.json,
quality_report.json) — промежуточное представление перед загрузкой в
PostgreSQL. Загрузка в БД — отдельный шаг (load_to_postgres.py),
не запускается автоматически.

Не удаляет и не перезаписывает исходный xlsx. Идемпотентен: повторный
запуск на том же файле даёт тот же результат (коды работ выводятся из
порядка строк в файле, не из состояния БД).
"""
import calendar
import json
import re
import sys
from collections import defaultdict
from datetime import date as date_cls, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

DEFAULT_YEAR = 2026  # объект должен завершиться в 2026 — см. РЕШЕНИЯ_v1.1.md
MONTHS_RU = {
    "ЯНВАРЬ": 1, "ФЕВРАЛЬ": 2, "МАРТ": 3, "АПРЕЛЬ": 4, "МАЙ": 5, "ИЮНЬ": 6,
    "ИЮЛЬ": 7, "АВГУСТ": 8, "СЕНТЯБРЬ": 9, "ОКТЯБРЬ": 10, "НОЯБРЬ": 11, "ДЕКАБРЬ": 12,
}

# ТЗ 8.1: участок/локация — закрытого справочника в самом Excel нет
# (колонки "Участок" не существует), код участка встречается как подстрока
# внутри "Наименование работ" (например "...УТ13", "...УУСА1"). Регэксп
# ищет ПЕРВОЕ совпадение по закрытому списку шифров участков — шифр,
# ОБЯЗАТЕЛЬНО сразу за которым идёт цифра. Бытовые русские слова капсом
# (ПАВИЛЬОНОВ, КАМЕРЫ, СКЛАДЕ и т.п.), встречающиеся в тексте наименования
# работы, участком не являются и намеренно не распознаются — они давали
# ложные срабатывания (see docs/REENGINEERING_LOG.md, находка "ПАВИЛЬОНОВ"
# вместо УУСА1 у TM35-IKS-001/002).
LOCATION_RE = re.compile(
    r"\b(УУСА\s?\d+(?:\.\d+)?|УУТЭ\s?\d+(?:\.\d+)?|УТП\s?\d+(?:\.\d+)?|УТ\s?\d+(?:\.\d+)?|"
    r"УП\s?\d+(?:\.\d+)?|ОПВ\s?\d+(?:\.\d+)?|ОПН\s?\d+(?:\.\d+)?|КР\s?\d+(?:\.\d+)?|"
    r"КМ\s?\d+(?:\.\d+)?|КД[\s-]?\d+(?:\.\d+)?|ТЭЦ)\b",
    re.IGNORECASE,
)


def guess_location(name):
    if not name:
        return None
    m = LOCATION_RE.search(name)
    return m.group(1).upper().replace(" ", "") if m else None


def parse_baseline_month(text):
    """
    "Месяц выполнения работ" (только на основном листе, не на всех строках
    заполнено) -> дата последнего дня ПОСЛЕДНЕГО упомянутого месяца, как
    временный baseline. "Август-Сентябрь-Октябрь" -> конец октября (более
    консервативная, чем первая, оценка срока). Возвращает None, если поле
    пустое или не содержит распознаваемого названия месяца.
    """
    if not text:
        return None
    found = [MONTHS_RU[m] for m in MONTHS_RU if re.search(m, str(text), re.IGNORECASE)]
    if not found:
        return None
    month = max(found)
    last_day = calendar.monthrange(DEFAULT_YEAR, month)[1]
    return f"{DEFAULT_YEAR:04d}-{month:02d}-{last_day:02d}"


def _is_weekend(d):
    return d.weekday() >= 5  # 5=Сб, 6=Вс


def _count_blocks(dates):
    """Число сплошных блоков плановых дней (выходные не считаются разрывом).
    См. docs/MATRIX_DIAGNOSTICS.md — тест 1/2."""
    dates = sorted(dates)
    blocks = 1
    for i in range(1, len(dates)):
        gap_days = (dates[i] - dates[i - 1]).days
        if gap_days <= 1:
            continue
        all_weekend = all(_is_weekend(dates[i - 1] + timedelta(days=k)) for k in range(1, gap_days))
        if not all_weekend:
            blocks += 1
    return blocks


def compute_matrix_baselines(daily_records):
    """
    По итогам диагностики (docs/MATRIX_DIAGNOSTICS.md, все 5 тестов
    согласованно "за график"): календарная матрица План интерпретируется
    как график работ, не ресурсная разнарядка. baseline для работы = первый/
    последний день с planned_crew > 0 в daily_progress (excel_import),
    точнее, чем текстовое "Месяц выполнения работ" (дневная, не месячная
    точность). Уверенность — high при 1 блоке (95% случаев), medium при
    2+ (возможна пауза в работе, окно не такое надёжное).

    daily_records: список словарей parse_excel.py (work_code, date, kind,
    value, ...) — те же, что уходят в daily_progress.json.
    Возвращает {work_code: {"plan_start":, "plan_finish":, "source":, "confidence":}}.
    """
    by_work = defaultdict(list)
    for r in daily_records:
        if r["kind"] != "plan":
            continue
        v = r["value"]
        if isinstance(v, (int, float)) and v > 0:
            by_work[r["work_code"]].append(date_cls.fromisoformat(r["date"]))

    result = {}
    for code, dates in by_work.items():
        dates = sorted(set(dates))
        blocks = _count_blocks(dates)
        result[code] = {
            "plan_start": dates[0].isoformat(),
            "plan_finish": dates[-1].isoformat(),
            "source": "matrix_schedule",
            "confidence": "high" if blocks == 1 else "medium",
        }
    return result


def parse_crew_number(crew_raw):
    """"Кол-во чел." из Excel — часто "4", иногда "4-5" (диапазон, берём
    верхнюю границу — консервативнее для расчёта потребности), иногда
    "Подряд"/пусто (не число). Не угадываем — просто None, если не парсится."""
    if crew_raw is None:
        return None
    if isinstance(crew_raw, (int, float)):
        return int(crew_raw)
    s = str(crew_raw).strip()
    m = re.match(r"^(\d+)\s*-\s*(\d+)$", s)
    if m:
        return int(m.group(2))
    m = re.match(r"^(\d+)$", s)
    if m:
        return int(m.group(1))
    return None


def derive_status(fact_pct):
    # Из Excel известен только "% вып." (снимок на дату выгрузки), не
    # статус приёмки — намеренно НЕ помечаем "closed"/"accepted"/
    # "submitted" (ТЗ 11.4.5: 100% физически ≠ принято/закрыто).
    if fact_pct is None:
        return "not_started"
    if fact_pct >= 100:
        return "done_physically"
    if fact_pct > 0:
        return "in_progress"
    return "not_started"


DAY_COMMENT_TYPE_RULES = [
    (re.compile(r"дожд|осадк|ветр|темпера", re.IGNORECASE), "weather"),
    (re.compile(r"гсм|топлив", re.IGNORECASE), "fuel"),
    (re.compile(r"техник", re.IGNORECASE), "equipment"),
]


def guess_blocker_type(text):
    for pattern, blocker_type in DAY_COMMENT_TYPE_RULES:
        if pattern.search(text):
            return blocker_type
    return "material"  # закрытый enum без "прочее" — материал как консервативный дефолт

AUX_SECTION_TITLE = "работы неучтенные в основном графике"

# По одному конфигу на лист: sheet-level source, и куда смотреть за метаколонками.
# Ищем колонку по вхождению одного из паттернов в текст заголовка (после
# нормализации пробелов), а не по фиксированной букве — см. skill.
SHEET_CONFIGS = {
    "График завершения работ по ТМ35": {
        "source": "main",
        "fields": {
            "seq_no": ["№ п/п"],
            "name": ["наименование работ"],
            "unit": ["ед. изм", "ед.изм"],
            "volume": ["объем работ", "объём работ"],
            "crew": ["кол-во чел"],
            "pct": ["% вып"],
            "equipment": ["вид техники"],
            "note": ["примечания"],
            "month_hint": ["месяц выполнения"],
        },
    },
    "График по замечаниям ИКС": {
        "source": "iks",
        "fields": {
            "seq_no": ["№ п/п"],
            "name": ["наименование работ"],
            "unit": ["ед. изм", "ед.изм"],
            "volume": ["объем работ", "объём работ"],
            "crew": ["кол-во чел"],
            "pct": ["% вып"],
            "equipment": ["вид техники"],
        },
    },
    "График по замечаниям РСК": {
        "source": "rsk",
        "fields": {
            "seq_no": ["№ п/п"],
            "registry_no": ["порядковый номер в реестре"],
            "content": ["содержание работы"],
            "name": ["наименование работ"],
            "unit": ["ед. изм", "ед.изм"],
            "volume": ["объем работ", "объём работ"],
            "crew": ["кол-во чел"],
            "pct": ["% вып"],
            "equipment": ["вид техники"],
        },
    },
}


def norm_header(text):
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip().lower()


def merged_anchor_value(ws, row, col):
    """Значение ячейки; если ячейка — часть merged-диапазона, значение из его top-left."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return ws.cell(row=mr.min_row, column=mr.min_col).value
    return None


def find_header_columns(ws, field_patterns, header_row=2, search_cols=20):
    """Находит номер колонки для каждого поля по тексту заголовка (row2)."""
    found = {}
    for col in range(1, search_cols + 1):
        h = norm_header(merged_anchor_value(ws, header_row, col))
        if not h:
            continue
        for field, patterns in field_patterns.items():
            if field in found:
                continue
            if any(p in h for p in patterns):
                found[field] = col
    missing = [f for f in field_patterns if f not in found]
    return found, missing


def build_calendar_columns(ws, header_row=2, day_row=3, subhdr_row=4, max_col=None):
    """
    Проходит по строкам 2-4 и строит список календарных колонок:
    [{"col": int, "kind": "plan"|"fact", "date": (year, month, day)}, ...]
    Плюс список колонок-разрывов (выходные, без День/План/Факт заголовка)
    в диапазоне календарной части листа.
    """
    max_col = max_col or ws.max_column
    # первая колонка календаря = первая колонка с непустым row2, содержащим
    # название месяца
    month_ranges = []  # (start_col, end_col, month_num)
    for mr in ws.merged_cells.ranges:
        if mr.min_row == header_row and mr.max_row == header_row:
            v = ws.cell(row=mr.min_row, column=mr.min_col).value
            if isinstance(v, str) and v.strip().upper() in MONTHS_RU:
                month_ranges.append((mr.min_col, mr.max_col, MONTHS_RU[v.strip().upper()]))
    if not month_ranges:
        return [], (None, None)
    month_ranges.sort()
    calendar_start = month_ranges[0][0]
    calendar_end = max(r[1] for r in month_ranges)

    def month_for_col(c):
        for start, end, m in month_ranges:
            if start <= c <= end:
                return m
        return None

    columns = []
    gap_columns = []
    col = calendar_start
    while col <= calendar_end:
        day_val = ws.cell(row=day_row, column=col).value
        if isinstance(day_val, (int, float)):
            month = month_for_col(col)
            day = int(day_val)
            columns.append({"col": col, "kind": "plan", "year": DEFAULT_YEAR, "month": month, "day": day})
            fact_col = col + 1
            fact_hdr = norm_header(ws.cell(row=subhdr_row, column=fact_col).value)
            if "факт" in fact_hdr:
                columns.append({"col": fact_col, "kind": "fact", "year": DEFAULT_YEAR, "month": month, "day": day})
                col = fact_col + 1
            else:
                col += 1
        else:
            gap_columns.append(col)
            col += 1
    return columns, gap_columns


def iter_data_rows(ws, start_row, meta_col_max):
    """Строки данных, начиная с start_row, до первого 3-кратного пустого разрыва в колонке A."""
    empty_streak = 0
    r = start_row
    while r <= ws.max_row:
        a_val = ws.cell(row=r, column=1).value
        row_has_any = any(
            ws.cell(row=r, column=c).value is not None for c in range(1, meta_col_max + 1)
        )
        if a_val is None and not row_has_any:
            empty_streak += 1
            if empty_streak >= 3:
                break
        else:
            empty_streak = 0
            yield r
        r += 1


def is_section_header_row(ws, row, name_col):
    a = ws.cell(row=row, column=1).value
    name = ws.cell(row=row, column=name_col).value if name_col else None
    if a is None:
        return False
    if isinstance(a, str) and not a.strip().isdigit():
        # текст в колонке A, остальное метаполе (имя работы) пусто
        if name is None:
            return True
    return False


def is_totals_row(ws, row):
    a = ws.cell(row=row, column=1).value
    return isinstance(a, str) and "итого" in a.lower()


def normalize_numeric_or_flag(raw):
    if raw is None:
        return None, None
    if isinstance(raw, (int, float)):
        return float(raw), str(raw)
    s = str(raw).strip()
    if s == "":
        return None, None
    m = re.match(r"^(\d+(?:[.,]\d+)?)\s*шт\.?$", s, re.IGNORECASE)
    if m:
        # физический объём в штуках, не процент
        return None, s
    m = re.match(r"^(\d+(?:[.,]\d+)?)$", s)
    if m:
        return float(m.group(1).replace(",", ".")), s
    return None, s


def parse_sheet(ws, sheet_name, cfg, code_prefix, quality_issues):
    fields, missing = find_header_columns(ws, cfg["fields"])
    if "name" not in fields:
        raise RuntimeError(f"[{sheet_name}] не найдена колонка «Наименование работ» — проверить шапку (см. skill)")

    calendar_cols, gap_cols = build_calendar_columns(ws)
    meta_col_max = max(fields.values(), default=9) + 1

    works = []
    daily = []
    seq = 0
    current_section = None
    started = False

    name_col = fields["name"]

    for row in iter_data_rows(ws, start_row=5, meta_col_max=meta_col_max):
        if is_totals_row(ws, row):
            # итоговая строка "Итого кол-во человек" — не работа, обрабатывается отдельно
            continue
        if is_section_header_row(ws, row, name_col):
            current_section = str(ws.cell(row=row, column=1).value).strip()
            started = True
            continue
        name_val = ws.cell(row=row, column=name_col).value
        if name_val is None or (isinstance(name_val, str) and not name_val.strip()):
            # "Наименование работ" пусто — на листе РСК встречается отдельно от
            # "Содержание работы" (описание дефекта), которое может быть заполнено,
            # когда наименование не указано. Не терять строку молча.
            content_col = fields.get("content")
            fallback = ws.cell(row=row, column=content_col).value if content_col else None
            if fallback is None or (isinstance(fallback, str) and not fallback.strip()):
                continue
            name_val = fallback
            row_quality_notes = [
                f"«Наименование работ» пусто ({get_column_letter(name_col)}{row}), "
                f"использовано «Содержание работы» — сверить с ПТО ИКС"
            ]
            quality_issues.append({
                "sheet": sheet_name, "row": row, "type": "name_missing_used_content_fallback",
                "cell": f"{get_column_letter(name_col)}{row}",
            })
        else:
            row_quality_notes = []

        seq += 1
        work_source = cfg["source"]
        if current_section and AUX_SECTION_TITLE in current_section.lower():
            work_source = "aux"

        code = f"TM35-{code_prefix[work_source]}-{seq:03d}"

        crew_raw = ws.cell(row=row, column=fields["crew"]).value if "crew" in fields else None
        executor_type = "own_forces"
        subcontractor_note = None
        if isinstance(crew_raw, str) and "подряд" in crew_raw.lower():
            executor_type = "subcontract"
            note_col = fields.get("note") or fields.get("equipment")
            subcontractor_note = ws.cell(row=row, column=note_col).value if note_col else None

        pct_val = ws.cell(row=row, column=fields["pct"]).value if "pct" in fields else None
        pct_norm, pct_raw = normalize_numeric_or_flag(pct_val)
        if pct_val is not None and pct_norm is None and pct_raw:
            quality_issues.append({
                "sheet": sheet_name, "row": row, "type": "pct_not_numeric",
                "value": pct_raw, "cell": f"{get_column_letter(fields['pct'])}{row}",
                "work_code": code,
            })
            row_quality_notes.append(
                f"«% вып.» = {pct_raw!r} не число ({get_column_letter(fields['pct'])}{row}) — "
                f"похоже на физический объём в штуках, не процент; не нормализовано автоматически"
            )

        volume_val = ws.cell(row=row, column=fields["volume"]).value if "volume" in fields else None
        volume_norm, volume_raw = normalize_numeric_or_flag(volume_val)
        if volume_val is not None and volume_norm is None and volume_raw:
            quality_issues.append({
                "sheet": sheet_name, "row": row, "type": "volume_not_numeric",
                "value": volume_raw, "cell": f"{get_column_letter(fields['volume'])}{row}",
                "work_code": code,
            })
            row_quality_notes.append(
                f"«Объём работ» = {volume_raw!r} не число ({get_column_letter(fields['volume'])}{row}) — "
                f"не нормализовано автоматически"
            )

        name_str = str(name_val).strip()
        work_rec = {
            "code": code,
            "source": work_source,
            "section": current_section,
            "name": name_str,
            "location": guess_location(name_str),
            "unit": ws.cell(row=row, column=fields["unit"]).value if "unit" in fields else None,
            "volume": volume_norm,
            "volume_raw": volume_raw,
            "crew_raw": crew_raw,
            "executor_type": executor_type,
            "status": derive_status(pct_norm),
            "baseline_finish": parse_baseline_month(
                ws.cell(row=row, column=fields["month_hint"]).value if "month_hint" in fields else None
            ),
            "baseline_crew": parse_crew_number(crew_raw),
            "equipment": ws.cell(row=row, column=fields["equipment"]).value if "equipment" in fields else None,
            "comment": ws.cell(row=row, column=fields["note"]).value if "note" in fields else None,
            "subcontractor_note": subcontractor_note,
            "fact_pct": pct_norm,
            "fact_pct_raw": pct_raw,
            "source_row_ref": f"{sheet_name}!{row}",
            "data_quality_flag": "needs_review" if row_quality_notes else "ok",
            "data_quality_note": "; ".join(row_quality_notes) or None,
        }
        works.append(work_rec)

        # календарные значения + комментарии по этой строке
        for cc in calendar_cols:
            cell = ws.cell(row=row, column=cc["col"])
            comment_text = cell.comment.text.strip() if cell.comment else None
            if cell.value is None and comment_text is None:
                continue
            try:
                iso_date = f"{cc['year']:04d}-{cc['month']:02d}-{cc['day']:02d}"
            except (TypeError, ValueError):
                # Дата не определяется однозначно (месяц не резолвится для этой
                # колонки) — валидного `date` для NOT NULL колонки daily_progress.date
                # нет, вставлять строку с угаданной датой нельзя (РЕШЕНИЯ_v1.1:
                # "не угадывать"). Уходит в import_unresolved_cell, не в daily_progress.
                quality_issues.append({
                    "sheet": sheet_name, "row": row, "type": "bad_calendar_date",
                    "cell": f"{get_column_letter(cc['col'])}{row}", "raw": cc,
                    "work_code": code, "value": cell.value,
                })
                continue
            daily.append({
                "work_code": code,
                "date": iso_date,
                "kind": cc["kind"],   # plan | fact
                "value": cell.value,
                "comment": comment_text,
                "cell": f"{get_column_letter(cc['col'])}{row}",
                "data_quality_flag": "ok",
                "data_quality_note": None,
            })

    # день-уровневые комментарии на строке 3 (заголовок дня) — не привязаны к работе.
    # Проверяем обе колонки пары (План И Факт) — комментарий встречается на любой
    # из них, дедуп по (дата, текст) на случай, если стоит на обеих.
    day_level_comments = []
    seen_day_comments = set()
    for cc in calendar_cols:
        cell = ws.cell(row=3, column=cc["col"])
        if cell.comment:
            date_str = f"{cc['year']:04d}-{cc['month']:02d}-{cc['day']:02d}"
            text = cell.comment.text.strip()
            key = (date_str, text)
            if key in seen_day_comments:
                continue
            seen_day_comments.add(key)
            day_level_comments.append({
                "sheet": sheet_name,
                "date": date_str,
                "comment": text,
                "blocker_type": guess_blocker_type(text),
            })

    return works, daily, day_level_comments


def main(xlsx_path, out_dir):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    code_prefix = {"main": "MAIN", "iks": "IKS", "rsk": "RSK", "aux": "AUX"}

    all_works, all_daily, all_day_comments = [], [], []
    quality_issues = []

    for sheet_name, cfg in SHEET_CONFIGS.items():
        if sheet_name not in wb.sheetnames:
            quality_issues.append({"sheet": sheet_name, "type": "sheet_missing"})
            continue
        ws = wb[sheet_name]
        works, daily, day_comments = parse_sheet(ws, sheet_name, cfg, code_prefix, quality_issues)
        all_works.extend(works)
        all_daily.extend(daily)
        all_day_comments.extend(day_comments)

    # Matrix-baseline (docs/MATRIX_DIAGNOSTICS.md) — точнее текстового
    # "Месяц выполнения работ", подменяет его там, где есть календарные
    # данные; иначе явный source='text_month_only'/'no_data', не молчим.
    matrix_baselines = compute_matrix_baselines(all_daily)
    for w in all_works:
        mb = matrix_baselines.get(w["code"])
        if mb:
            w["baseline_start"] = mb["plan_start"]
            w["baseline_finish"] = mb["plan_finish"]  # переопределяет текстовый месяц точной датой
            w["baseline_source"] = mb["source"]
            w["baseline_confidence"] = mb["confidence"]
        elif w.get("baseline_finish"):
            w["baseline_start"] = None
            w["baseline_source"] = "text_month_only"
            w["baseline_confidence"] = "low"
        else:
            w["baseline_start"] = None
            w["baseline_source"] = "no_data"
            w["baseline_confidence"] = "none"

    (out_dir / "works.json").write_text(
        json.dumps(all_works, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out_dir / "daily_progress.json").write_text(
        json.dumps(all_daily, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out_dir / "day_level_comments.json").write_text(
        json.dumps(all_day_comments, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out_dir / "quality_report.json").write_text(
        json.dumps(quality_issues, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    by_source = defaultdict(int)
    for w in all_works:
        by_source[w["source"]] += 1

    print(f"Работ разобрано: {len(all_works)}  {dict(by_source)}")
    print(f"Записей план/факт: {len(all_daily)}")
    print(f"День-уровневых комментариев: {len(all_day_comments)}")
    print(f"Замечаний качества данных: {len(quality_issues)}")
    print(f"Результат: {out_dir}/")


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else \
        "/home/oleg/Documents/TM-35/График работ до завершения стройка_ИКС_РСК 13.08 (2).xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "/home/oleg/Documents/TM-35/import/output"
    main(xlsx, out)
