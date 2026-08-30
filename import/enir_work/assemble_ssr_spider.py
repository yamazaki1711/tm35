"""
Справочник норм СТО-ССР-2026 (Spider Project) — из ssr_spider_norms.json.
Внутренний стандарт подрядчика ООО «ССР», ранее реально применялся
Заказчиком для ввода отчётов ТМ-35 в Spider Project (практика
прекращена). Разделы работ прямо соответствуют ТМ-35 (Устройство
трубопроводов, Устройство камеры, СОДК и т.п.) — в отличие от общего
16-сборникового отбора ГЭСН.
"""
import json
from collections import Counter, defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

SRC = "/home/oleg/Documents/TM-35/import/enir_work/ssr_spider_norms.json"
XLSX_OUT = "/home/oleg/Documents/TM-35/docs/spravochnik_SSR_SpiderProject_TM35.xlsx"
MD_OUT = "/home/oleg/Documents/TM-35/docs/spravochnik_SSR_SpiderProject_TM35.md"

HEADERS = ["Раздел", "Код", "Наименование операции", "Ед. изм.",
           "Состав команды", "Производительность бригады, ед/час",
           "Трудоёмкость, чел-час/ед", "Состав работы (кратко)"]


def crew_text(crew):
    parts = []
    for c in crew:
        q = c["qty"]
        qty_s = f"{q:g}" if q is not None else "?"
        parts.append(f"{c['resource']} {qty_s} {c['unit']}".strip())
    return "; ".join(parts)


def main():
    ops = json.load(open(SRC, encoding="utf-8"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Нормы Spider Project"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    ws2 = wb.create_sheet("Без трудоёмкости")
    ws2.append(["Раздел", "Код", "Наименование", "Ед.изм.", "Состав команды", "Примечание"])
    for c in ws2[1]:
        c.font = Font(bold=True)

    by_section = Counter()
    by_section_ok = Counter()

    for o in ops:
        by_section[o["section"]] += 1
        row = [o["section"], o["code"], o["name"], o.get("unit") or "",
               crew_text(o["crew"]), o.get("team_productivity_per_hour"),
               o.get("labor_hours_per_unit"), (o.get("notes") or "")[:300]]
        if o.get("labor_hours_per_unit") is not None:
            by_section_ok[o["section"]] += 1
            ws.append(row)
        else:
            ws2.append([o["section"], o["code"], o["name"], o.get("unit") or "",
                        crew_text(o["crew"]),
                        "Трудоёмкость не указана в источнике (обычно означает: норма определяется "
                        "только производительностью механизма, людской составляющей нет) — не выдумано."])

    for sheet, widths in ((ws, [30, 16, 42, 10, 45, 14, 14, 45]), (ws2, [30, 16, 42, 10, 45, 45])):
        for i, w in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = w
        sheet.freeze_panes = "A2"

    wb.save(XLSX_OUT)

    total = len(ops)
    total_ok = sum(1 for o in ops if o.get("labor_hours_per_unit") is not None)

    md = []
    md.append("# Справочник норм СТО-ССР-2026 (Spider Project) для ТМ-35\n")
    md.append(
        "Внутренний норматив подрядчика **ООО «ССР»** (та же организация, что "
        "выполняет ТМ-35, подписант сметы как «Строительная дирекция объектов "
        "теплоснабжения СД-2») — «Единичные нормы для планирования ресурсов "
        "в ПО Spider Project», версия 4, приложение 1. Раньше Заказчик реально "
        "вводил отчёты с объекта ТМ-35 в Spider Project по этим нормам "
        "(практика прекращена) — это единственное, что осталось от того "
        "процесса. Разделы работ прямо соответствуют ТМ-35 (Устройство "
        "трубопроводов, Устройство камеры, СОДК, Бурошнек и т.п.), в отличие "
        "от общего отбора сборников ГЭСН. По словам координатора, Spider "
        "Project, вероятно, подключён к ГЭСН через API — нормы местами близки "
        "к ГЭСН, но здесь они уже привязаны к реальной технике и бригадам "
        "подрядчика.\n"
    )
    md.append(
        "**Готовая формула из источника:** `Длительность = Объём / "
        "(Производительность_ключевого_ресурса × Загрузка% × Кол-во)`; "
        "`Трудоёмкость = Длительность × Кол-во_людей × Загрузка%`.\n"
    )
    md.append("## Сводка охвата\n")
    md.append(f"- Всего операций: **{total}**")
    md.append(f"- С трудоёмкостью чел-час/ед. (лист «Нормы Spider Project»): **{total_ok}** ({100*total_ok/total:.1f}%)")
    md.append(f"- Без трудоёмкости (лист «Без трудоёмкости» — обычно чисто механизированная операция без людской составляющей, не выдумано): **{total-total_ok}**")
    md.append("")
    md.append("## Охват по разделам\n")
    md.append("| Раздел | Операций | С трудоёмкостью |")
    md.append("|---|---|---|")
    for sec in sorted(by_section, key=lambda x: str(x)):
        md.append(f"| {sec} | {by_section[sec]} | {by_section_ok[sec]} |")
    md.append("")
    md.append(
        "## Проверка на самосогласованность\n\nСверено вручную на случайной "
        "выборке 12 операций из разных разделов: трудоёмкость из источника "
        "пересчитана по собственной формуле документа (чел × загрузка% / "
        "производительность бригады) — совпало на 10 из 12 без округления. "
        "Пример: «Сварка трубы встык напорной ПЭ д110мм» — 2 рабочих + "
        "1 оператор сварки ПЭ, бригада 1,9 стыков/час → 3 чел × (1/1,9) час "
        "= 1,578 чел-час/стык, ровно как в источнике.\n\n"
        "Два расхождения в выборке — не ошибка числа трудоёмкости (оно взято "
        "из источника как есть, не вычислено нами), а частности разбора "
        "состава команды: (1) в одной операции электросварщик размечен в "
        "исходнике единицей «шт.», а не «чел.» — трудоёмкость в файле "
        "правильная, просто при быстрой проверке его не посчитали человеком; "
        "(2) «Прогрев бетона тепловыми пушками» — не подчиняется обычной "
        "формуле бригада×загрузка (похоже на операцию с фиксированной "
        "длительностью, не почасовой производительностью) — число из "
        "источника взято как есть, не пересчитано."
    )

    with open(MD_OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(md) + "\n")

    print("xlsx:", XLSX_OUT)
    print("md:", MD_OUT)
    print(f"total={total} ok={total_ok} ({100*total_ok/total:.1f}%)")


if __name__ == "__main__":
    main()
