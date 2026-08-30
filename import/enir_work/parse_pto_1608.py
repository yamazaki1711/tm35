"""
Разбор нового снимка Excel ПТО (16.08) — по правилам .claude/skills/tm35-excel.
Один лист, метаколонки A..I. Цель — структурированный список работ
(наименование/ед.изм./объём) для проверки, сколько теперь реально
нормируемо (физическая единица, не 'комп.').
"""
import json

import openpyxl

SRC = "/home/oleg/Downloads/График работ до завершения стройка_ИКС_РСК 16.08.xlsx"
OUT = "/home/oleg/Documents/TM-35/import/enir_work/pto_1608.json"

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb.active

items = []
current_section = None
empty_streak = 0

for r in range(5, ws.max_row + 1):
    a = ws.cell(row=r, column=1).value
    b = ws.cell(row=r, column=2).value
    if a is None and b is None:
        empty_streak += 1
        if empty_streak >= 3:
            break
        continue
    empty_streak = 0

    if isinstance(a, str) and b is None:
        current_section = a.strip()
        continue

    if isinstance(a, (int, float)) and isinstance(b, str):
        unit = ws.cell(row=r, column=3).value
        vol = ws.cell(row=r, column=4).value
        crew = ws.cell(row=r, column=5).value
        pct = ws.cell(row=r, column=7).value
        items.append({
            "n": int(a), "section": current_section, "name": b.strip(),
            "unit": (unit or "").strip() if isinstance(unit, str) else unit,
            "qty": vol, "crew_raw": crew, "pct": pct,
        })

json.dump(items, open(OUT, "w", encoding="utf-8"), ensure_ascii=False, indent=1)
print(f"Позиций: {len(items)}")

from collections import Counter
units = Counter((i["unit"] or "").strip().lower() for i in items)
print("Единицы измерения:", units.most_common(15))

physical = {"шт", "м3", "м2", "м", "м.п.", "п.м.", "т", "мп"}
n_physical = sum(1 for i in items if (i["unit"] or "").strip().lower() in physical and i.get("qty") not in (None, ""))
print(f"С физической единицей и объёмом: {n_physical} из {len(items)}")
