"""
Шаг 4: собрать итоговый справочник из matched.json в
docs/spravochnik_norm_TM35.xlsx (таблица) и .md (тот же список +
явная сводка охвата — сколько позиций/объёма сметы сопоставлено).
"""
import json
from collections import Counter

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

SRC = "/home/oleg/Documents/TM-35/import/enir_work/matched.json"
XLSX_OUT = "/home/oleg/Documents/TM-35/docs/spravochnik_norm_TM35.xlsx"
MD_OUT = "/home/oleg/Documents/TM-35/docs/spravochnik_norm_TM35.md"

HEADERS = [
    "№ п/п (смета)", "Раздел РД", "Наименование работы", "Ед.изм.",
    "Объём по смете", "Сборник/§ ЕНиР", "Норма (фраза из ЕНиР, чел-час/ед.)",
    "Состав звена (специализация, разряд)", "Уверенность сопоставления",
    "Примечание",
]

STATUS_FILL = {
    "точное": PatternFill("solid", fgColor="C6EFCE"),
    "эвристическое": PatternFill("solid", fgColor="FFEB9C"),
    "не найдено": PatternFill("solid", fgColor="FFC7CE"),
}


def build_note(r):
    if r["match_status"] == "не найдено":
        return "Норма ЕНиР не найдена по доступным сборникам — не выдумано, требует ручного подбора либо другого источника (например, ГЭСН/ФЕР)."
    if r["match_status"] == "эвристическое":
        return "Сопоставление по совпадению ключевых слов — не проверено вручную, перед использованием сверить с текстом §."
    return "Совпадение по ключевым словам и единице измерения — рекомендуется точечная проверка перед вводом в расчёт."


def main():
    results = json.load(open(SRC, encoding="utf-8"))
    results.sort(key=lambda r: r["n"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Справочник"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    for r in results:
        crew = r.get("enir_crew_raw") or ""
        row = [
            r["n"],
            f"{r['section_num']}. {r['section_name']}" if r.get("section_num") else "",
            r["name"],
            r.get("unit") or "",
            r.get("qty"),
            f"{r['enir_code']} ({r['enir_sbornik']})" if r.get("enir_code") else "",
            (r.get("enir_unit_phrase") or "")[:200],
            crew[:300],
            r["match_status"],
            build_note(r),
        ]
        ws.append(row)
        fill = STATUS_FILL.get(r["match_status"])
        if fill:
            ws.cell(row=ws.max_row, column=9).fill = fill

    widths = [8, 24, 42, 8, 12, 22, 30, 34, 14, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    wb.save(XLSX_OUT)

    # --- сводка охвата ---
    status_counts = Counter(r["match_status"] for r in results)
    total = len(results)

    by_sbornik = Counter(r["enir_sbornik"] for r in results if r.get("enir_sbornik"))

    md = []
    md.append("# Справочник норм ЕНиР для ТМ-35\n")
    md.append(
        "Сопоставление 512 позиций ведомости работ по РД (\"Смета контракта "
        "-ред.ДС№16 от 28.05.25.xlsx\", Google Drive объекта, папка "
        "«Хабаровск ТМ-35» → 09_СМЕТЫ И НК) с параграфами норм ЕНиР "
        "(Сборники Е1–Е40, `/home/oleg/Documents/TM-35/ЕНиР/`). "
        "Метод — детерминированное сопоставление по пересечению ключевых "
        "слов наименования и единице измерения, без угадывания LLM. "
        "Полная таблица — `spravochnik_norm_TM35.xlsx`.\n"
    )
    md.append("## Сводка охвата\n")
    md.append(f"- Всего позиций ведомости: **{total}**")
    for status in ("точное", "эвристическое", "не найдено"):
        n = status_counts.get(status, 0)
        md.append(f"- {status}: **{n}** ({100*n/total:.1f}%)")
    md.append("")
    md.append(
        "**Точное/эвристическое — не значит проверено вручную.** Оба "
        "статуса — результат текстового сопоставления; координатору "
        "рекомендуется выборочно свериться с исходным § ЕНиР перед тем "
        "как закладывать конкретные числа в расчёт трудоёмкости."
    )
    md.append("")
    md.append("## Охват по сборникам ЕНиР (учтено в этой версии)\n")
    def ru_count(n):
        if n % 10 == 1 and n % 100 != 11:
            return "сопоставление"
        if 2 <= n % 10 <= 4 and not (12 <= n % 100 <= 14):
            return "сопоставления"
        return "сопоставлений"

    for name, n in sorted(by_sbornik.items()):
        md.append(f"- {name}: {n} {ru_count(n)}")
    md.append("")
    md.append("## Известные пробелы (честно, не скрыто)\n")
    md.append(
        "- В локальном архиве `ЕНиР/` отсутствуют сборники на монолитные "
        "и сборные железобетонные конструкции (стандартно — Е4) — раздел "
        "14 ведомости («Конструктивные решения: фундаменты...») почти "
        "без сопоставлений именно по этой причине, не по вине парсера."
    )
    md.append(
        "- Отдельного сборника ЕНиР на пусконаладочные работы или КИПиА "
        "в архиве нет (в стандартной линейке ЕНиР Е1–Е40 такого сборника "
        "не существует вообще) — эти позиции ведомости по определению "
        "не могут быть закрыты этим источником."
    )
    md.append(
        "- Не сопоставленные позиции (см. таблицу, статус «не найдено») "
        "не отброшены — присутствуют в xlsx с исходным наименованием и "
        "объёмом, требуют либо расширения перечня сборников, либо другого "
        "источника норм (ГЭСН/ФЕР)."
    )
    md.append("")
    md.append(
        "## Топ-20 несопоставленных позиций по объёму (для приоритизации)\n"
    )
    md.append(
        "Из ранжирования по объёму исключены позиции разделов 35 «Прочие» "
        "и 36 «Непредвиденные затраты», где единица измерения — «руб.» "
        "(авторский надзор, экспертиза, резерв на непредвиденные затраты "
        "и т.п.) — это не физические работы, для них норма ЕНиР "
        "в принципе не существует, не пробел справочника.\n"
    )
    unmatched = [
        r for r in results
        if r["match_status"] == "не найдено" and (r.get("unit") or "").strip().lower() != "руб."
    ]

    def qty_key(r):
        try:
            return float(r.get("qty") or 0)
        except (TypeError, ValueError):
            return 0.0

    unmatched.sort(key=qty_key, reverse=True)
    md.append("| № | Наименование | Ед.изм. | Объём |")
    md.append("|---|---|---|---|")
    for r in unmatched[:20]:
        md.append(f"| {r['n']} | {r['name']} | {r.get('unit') or ''} | {r.get('qty')} |")

    with open(MD_OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(md) + "\n")

    print("xlsx:", XLSX_OUT)
    print("md:", MD_OUT)
    print(dict(status_counts))


if __name__ == "__main__":
    main()
