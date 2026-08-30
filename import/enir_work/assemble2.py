"""
Итоговый справочник норм ЕНиР (прямая оцифровка, без привязки к смете
ТМ-35) — из enir_norms_v2.json. Одна строка = одна комбинация параметров
со своим числом Н.вр. Параграфы, где таблицу не удалось разобрать
автоматически (пустые исходники или нераспознанный формат) — отдельным
списком, без выдуманных чисел.
"""
import json
from collections import defaultdict

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

SRC = "/home/oleg/Documents/TM-35/import/enir_work/enir_norms_v2.json"
XLSX_OUT = "/home/oleg/Documents/TM-35/docs/spravochnik_ENiR_TM35.xlsx"
MD_OUT = "/home/oleg/Documents/TM-35/docs/spravochnik_ENiR_TM35.md"

HEADERS = ["Сборник", "§", "Наименование работы", "Условие/параметр",
           "Ед. изм.", "Норма времени, чел-час", "Состав звена", "Примечание"]


def sbornik_short(name):
    return name.replace("ЕНиР Сборник ", "").replace("ЕНиР ", "")


def main():
    rows = json.load(open(SRC, encoding="utf-8"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Нормы (по параметрам)"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    parsed = [r for r in rows if r["parsed"]]
    unparsed = [r for r in rows if not r["parsed"]]

    for r in parsed:
        ws.append([
            sbornik_short(r["sbornik"]),
            r["code"],
            r["title"],
            r.get("condition") or "",
            r.get("unit_phrase") or "",
            r.get("hours_per_unit"),
            r.get("crew_raw") or "",
            "",
        ])

    ws2 = wb.create_sheet("Не разобрано автоматически")
    ws2.append(["Сборник", "§", "Наименование работы", "Ед. изм.", "Состав звена", "Примечание"])
    for c in ws2[1]:
        c.font = Font(bold=True)
    for r in unparsed:
        ws2.append([
            sbornik_short(r["sbornik"]), r["code"], r["title"],
            r.get("unit_phrase") or "", r.get("crew_raw") or "",
            "Таблица не разобрана автоматически (пустая в источнике либо "
            "нераспознанный формат) — числа не выдуманы, см. первоисточник "
            "в ЕНиР/.",
        ])

    for sheet, widths in ((ws, [16, 12, 40, 40, 14, 14, 34, 30]), (ws2, [16, 12, 40, 12, 34, 45])):
        for i, w in enumerate(widths, 1):
            sheet.column_dimensions[get_column_letter(i)].width = w
        sheet.freeze_panes = "A2"

    wb.save(XLSX_OUT)

    # --- сводка ---
    by_sbornik_paragraphs = defaultdict(lambda: [0, 0])  # [всего §, с нормой]
    seen_codes = set()
    for r in rows:
        key = (r["sbornik"], r["code"])
        if key in seen_codes:
            continue
        seen_codes.add(key)
        by_sbornik_paragraphs[sbornik_short(r["sbornik"])][0] += 1
        if r["parsed"]:
            by_sbornik_paragraphs[sbornik_short(r["sbornik"])][1] += 1

    total_paragraphs = len(seen_codes)
    total_with_norm = sum(1 for k in seen_codes if any(
        rr["parsed"] for rr in rows if (rr["sbornik"], rr["code"]) == k
    ))
    # быстрее: пересчитать через множество распарсенных кодов
    parsed_codes = {(r["sbornik"], r["code"]) for r in parsed}
    total_with_norm = len(parsed_codes)

    md = []
    md.append("# Справочник норм ЕНиР — прямая оцифровка (16 сборников, релевантных ТМ-35)\n")
    md.append(
        "Прямое извлечение параграфов (§) из книг ЕНиР "
        "(`/home/oleg/Documents/TM-35/ЕНиР/`) — без привязки к какой-либо "
        "конкретной смете. Таблицы норм времени разобраны через реальную "
        "структуру HTML-таблиц (colspan/строки), не через плоский текст — "
        "числа сверены вручную по нескольким сборникам разного формата "
        "(Е22, Е5, Е26) на совпадение с оригиналом. Один параграф с "
        "параметрической нормой (диаметр/масса/способ и т.п.) "
        "**развёрнут в несколько строк** — по одной строке на комбинацию "
        "параметров со своим числом Н.вр., как и просил координатор.\n"
    )
    md.append("## Сводка охвата\n")
    md.append(f"- Всего параграфов (§) в 16 сборниках: **{total_paragraphs}**")
    md.append(
        f"- С автоматически разобранной нормой (лист «Нормы (по "
        f"параметрам)»): **{total_with_norm}** "
        f"({100*total_with_norm/total_paragraphs:.1f}%), это "
        f"**{len(parsed)}** строк-комбинаций параметров"
    )
    md.append(
        f"- Не разобрано автоматически (лист «Не разобрано автоматически» "
        f"— название/состав звена есть, числа нет, не придумано): "
        f"**{total_paragraphs - total_with_norm}** "
        f"({100*(total_paragraphs - total_with_norm)/total_paragraphs:.1f}%)"
    )
    md.append("")
    md.append(
        "**Почему не 100%.** Два разных случая, оба честно видны в "
        "листе «Не разобрано»: (1) формат таблицы не подошёл ни под один "
        "из распознанных парсером шаблонов; (2) в самом файле .doc ячейки "
        "с числами Н.вр. физически пустые — проверено прямо в HTML-"
        "исходнике (не артефакт конвертации), например весь "
        "`ЕНиР Сборник Е 9 Выпуск 2` — таблицы там присутствуют, но "
        "ячейки со значениями пустые в самом архиве."
    )
    md.append("")
    md.append("## Охват по сборникам\n")
    md.append("| Сборник | § всего | § с нормой | % |")
    md.append("|---|---|---|---|")
    for name in sorted(by_sbornik_paragraphs):
        total, _ = by_sbornik_paragraphs[name]
        with_norm = len({r["code"] for r in parsed if sbornik_short(r["sbornik"]) == name})
        pct = 100 * with_norm / total if total else 0
        md.append(f"| {name} | {total} | {with_norm} | {pct:.0f}% |")

    with open(MD_OUT, "w", encoding="utf-8") as f:
        f.write("\n".join(md) + "\n")

    print("xlsx:", XLSX_OUT)
    print("md:", MD_OUT)
    print(f"§ total={total_paragraphs} with_norm={total_with_norm} rows={len(parsed)}")


if __name__ == "__main__":
    main()
