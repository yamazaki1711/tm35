#!/usr/bin/env python3
"""
Batch-импорт графика работ ТМ-35 из Excel в нормализованные таблицы.

Структура файла и алгоритм разбора описаны в
.claude/skills/tm35-excel/SKILL.md — читать перед правкой этого файла.

Выход: JSON-файлы в import/output/ (works.json, daily_progress.json,
quality_report.json) — промежуточное представление перед загрузкой в
PostgreSQL. Загрузка в БД — отдельный шаг (load_to_postgres.py),
не запускается автоматически.

Не удаляет и не перезаписывает исходный xlsx. Идемпотентен: повторный
запуск на том же файле даёт тот же результат (коды работ выводятся из
порядка строк в файле, не из состояния БД).
"""
import calendar
import json
import re
import sys
from collections import defaultdict
from datetime import date as date_cls, timedelta
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter

DEFAULT_YEAR = 2026  # объект должен завершиться в 2026 — см. РЕШЕНИЯ_v1.1.md

# Найдено 29.08.2026 при разборе нового файла: часть ячеек содержит
# современные Excel "threaded comments", не старые VML-примечания —
# openpyxl отдаёт их текст обёрнутым в служебный текст самого Excel
# ("[Threaded comment]... Your application allows you to read this
# threaded comment... Comment:\n\t\t<текст>"), который иначе утекал бы
# как есть в БД (найдено на живом дашборде — /dashboard, «Стоп-факторы»
# показывал этот текст вместо «Осадки в виде дождя,отсутствие ГСМ»).
# Извлекаем только то, что после "Comment:" — остальное отбрасываем.
_THREADED_COMMENT_RE = re.compile(r"comment:\s*", re.IGNORECASE)


def clean_comment_text(raw):
    if raw is None:
        return None
    text = raw.strip()
    if not text:
        return None
    if text.startswith("[Threaded comment]"):
        parts = _THREADED_COMMENT_RE.split(text, maxsplit=1)
        if len(parts) == 2:
            text = parts[1].strip()
    return text or None


def clean_text_field(raw):
    """
    НАЙДЕНО 31.08.2026: \xa0 (неразрывный пробел) массово встречается не
    только в одиночных «пустых» ячейках календаря (см. build_calendar_columns),
    но и ВНУТРИ текстовых полей — 11 наименований работ, 58 примечаний,
    22 «вид техники», 3 ед.изм. В наименованиях он ломает точное сопоставление
    со старыми кодами БД в reconcile (пример: код TM35-MAIN-001 «...изоляции  УТ6»
    в БД против «...изоляции\xa0 УТ6» в новом файле — визуально одно и то же,
    программно разные строки, работа улетела бы в «новые» и потеряла историю).
    Заменяем на обычный пробел и схлопываем — не угадывание содержимого,
    только нормализация невидимого юникод-артефакта. Пусто после очистки -> None.
    """
    if raw is None or not isinstance(raw, str):
        return raw
    text = re.sub(r"\s+", " ", raw.replace("\xa0", " ")).strip()
    return text or None


MONTHS_RU = {
    "ЯНВАРЬ": 1, "ФЕВРАЛЬ": 2, "МАРТ": 3, "АПРЕЛЬ": 4, "МАЙ": 5, "ИЮНЬ": 6,
    "ИЮЛЬ": 7, "АВГУСТ": 8, "СЕНТЯБРЬ": 9, "ОКТЯБРЬ": 10, "НОЯБРЬ": 11, "ДЕКАБРЬ": 12,
}

# ТЗ 8.1: участок/локация — закрытого справочника в самом Excel нет
# (колонки "Участок" не существует), код участка встречается как подстрока
# внутри "Наименование работ" (например "...УТ13", "...УУСА1"). Регэксп
# ищет ПЕРВОЕ совпадение по закрытому списку шифров участков — шифр,
# ОБЯЗАТЕЛЬНО сразу за которым идёт цифра. Бытовые русские слова капсом
# (ПАВИЛЬОНОВ, КАМЕРЫ, СКЛАДЕ и т.п.), встречающиеся в тексте наименования
# работы, участком не являются и намеренно не распознаются — они давали
# ложные срабатывания (see docs/REENGINEERING_LOG.md, находка "ПАВИЛЬОНОВ"
# вместо УУСА1 у TM35-IKS-001/002).
LOCATION_RE = re.compile(
    r"\b(УУСА\s?\d+(?:\.\d+)?|УУТЭ\s?\d+(?:\.\d+)?|УТП\s?\d+(?:\.\d+)?|УТ\s?\d+(?:\.\d+)?|"
    r"УП\s?\d+(?:\.\d+)?|ОПВ\s?\d+(?:\.\d+)?|ОПН\s?\d+(?:\.\d+)?|КР\s?\d+(?:\.\d+)?|"
    r"КМ\s?\d+(?:\.\d+)?|КД[\s-]?\d+(?:\.\d+)?|ТЭЦ)\b",
    re.IGNORECASE,
)


def guess_location(name):
    if not name:
        return None
    m = LOCATION_RE.search(name)
    return m.group(1).upper().replace(" ", "") if m else None


def parse_baseline_month(text):
    """
    "Месяц выполнения работ" (только на основном листе, не на всех строках
    заполнено) -> дата последнего дня ПОСЛЕДНЕГО упомянутого месяца, как
    временный baseline. "Август-Сентябрь-Октябрь" -> конец октября (более
    консервативная, чем первая, оценка срока). Возвращает None, если поле
    пустое или не содержит распознаваемого названия месяца.
    """
    if not text:
        return None
    found = [MONTHS_RU[m] for m in MONTHS_RU if re.search(m, str(text), re.IGNORECASE)]
    if not found:
        return None
    month = max(found)
    last_day = calendar.monthrange(DEFAULT_YEAR, month)[1]
    return f"{DEFAULT_YEAR:04d}-{month:02d}-{last_day:02d}"


def _is_weekend(d):
    return d.weekday() >= 5  # 5=Сб, 6=Вс


def _count_blocks(dates):
    """Число сплошных блоков плановых дней (выходные не считаются разрывом).
    См. docs/MATRIX_DIAGNOSTICS.md — тест 1/2."""
    dates = sorted(dates)
    blocks = 1
    for i in range(1, len(dates)):
        gap_days = (dates[i] - dates[i - 1]).days
        if gap_days <= 1:
            continue
        all_weekend = all(_is_weekend(dates[i - 1] + timedelta(days=k)) for k in range(1, gap_days))
        if not all_weekend:
            blocks += 1
    return blocks


def compute_matrix_baselines(daily_records):
    """
    По итогам диагностики (docs/MATRIX_DIAGNOSTICS.md, все 5 тестов
    согласованно "за график"): календарная матрица План интерпретируется
    как график работ, не ресурсная разнарядка. baseline для работы = первый/
    последний день с planned_crew > 0 в daily_progress (excel_import),
    точнее, чем текстовое "Месяц выполнения работ" (дневная, не месячная
    точность). Уверенность — high при 1 блоке (95% случаев), medium при
    2+ (возможна пауза в работе, окно не такое надёжное).

    daily_records: список словарей parse_excel.py (work_code, date, kind,
    value, ...) — те же, что уходят в daily_progress.json.
    Возвращает {work_code: {"plan_start":, "plan_finish":, "source":, "confidence":}}.
    """
    by_work = defaultdict(list)
    for r in daily_records:
        if r["kind"] != "plan":
            continue
        v = r["value"]
        if isinstance(v, (int, float)) and v > 0:
            by_work[r["work_code"]].append(date_cls.fromisoformat(r["date"]))

    result = {}
    for code, dates in by_work.items():
        dates = sorted(set(dates))
        blocks = _count_blocks(dates)
        result[code] = {
            "plan_start": dates[0].isoformat(),
            "plan_finish": dates[-1].isoformat(),
            "source": "matrix_schedule",
            "confidence": "high" if blocks == 1 else "medium",
        }
    return result


def parse_crew_number(crew_raw):
    """"Кол-во чел." из Excel — часто "4", иногда "4-5" (диапазон, берём
    верхнюю границу — консервативнее для расчёта потребности), иногда
    "Подряд"/пусто (не число). Не угадываем — просто None, если не парсится."""
    if crew_raw is None:
        return None
    if isinstance(crew_raw, (int, float)):
        return int(crew_raw)
    s = str(crew_raw).strip()
    m = re.match(r"^(\d+)\s*-\s*(\d+)$", s)
    if m:
        return int(m.group(2))
    m = re.match(r"^(\d+)$", s)
    if m:
        return int(m.group(1))
    return None


def derive_status(fact_pct):
    # Из Excel известен только "% вып." (снимок на дату выгрузки), не
    # статус приёмки — намеренно НЕ помечаем "closed"/"accepted"/
    # "submitted" (ТЗ 11.4.5: 100% физически ≠ принято/закрыто).
    if fact_pct is None:
        return "not_started"
    if fact_pct >= 100:
        return "done_physically"
    if fact_pct > 0:
        return "in_progress"
    return "not_started"


DAY_COMMENT_TYPE_RULES = [
    (re.compile(r"дожд|осадк|ветр|темпера", re.IGNORECASE), "weather"),
    (re.compile(r"гсм|топлив", re.IGNORECASE), "fuel"),
    (re.compile(r"техник", re.IGNORECASE), "equipment"),
]


def guess_blocker_type(text):
    for pattern, blocker_type in DAY_COMMENT_TYPE_RULES:
        if pattern.search(text):
            return blocker_type
    return "material"  # закрытый enum без "прочее" — материал как консервативный дефолт

AUX_SECTION_TITLE = "работы неучтенные в основном графике"

# ВЕРСИЯ v3 (31.08.2026) — источник «Сводная Таблица по ТМ 35 от
# 30.08.26.xlsx» (файл идентичен по MD5 копии «27.08.26 (2)» — «27.08» в
# имени устаревшая датировка, содержимое актуальное, координатор
# подтвердил). Лист переименован с «СМР общий ГПР» на «График СМР» —
# только имя листа, структура (2 секции по заголовку в колонке A: «Работы
# по устранению выявленных недостатков при строительстве» = main,
# «Работы неучтенные в основном графике» = aux) не изменилась.
SHEET_CONFIGS = {
    "График СМР": {
        "source": "main",
        "fields": {
            "seq_no": ["№ п/п"],
            "name": ["наименование работ"],
            "unit": ["ед. изм", "ед.изм"],
            "volume": ["объем работ", "объём работ"],
            "crew": ["кол-во чел"],
            "pct": ["% вып"],
            "equipment": ["вид техники"],
            "note": ["примечания"],
            "month_hint": ["месяц выполнения"],
        },
    },
}


def norm_header(text):
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip().lower()


def merged_anchor_value(ws, row, col):
    """Значение ячейки; если ячейка — часть merged-диапазона, значение из его top-left."""
    cell = ws.cell(row=row, column=col)
    if cell.value is not None:
        return cell.value
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return ws.cell(row=mr.min_row, column=mr.min_col).value
    return None


def find_header_columns(ws, field_patterns, header_row=2, search_cols=20):
    """Находит номер колонки для каждого поля по тексту заголовка (row2)."""
    found = {}
    for col in range(1, search_cols + 1):
        h = norm_header(merged_anchor_value(ws, header_row, col))
        if not h:
            continue
        for field, patterns in field_patterns.items():
            if field in found:
                continue
            if any(p in h for p in patterns):
                found[field] = col
    missing = [f for f in field_patterns if f not in found]
    return found, missing


def build_calendar_columns(ws, header_row=2, day_row=3, subhdr_row=4, max_col=None):
    """
    Проходит по строкам 2-4 и строит список календарных колонок:
    [{"col": int, "kind": "plan"|"fact", "date": (year, month, day)}, ...]
    Плюс список колонок-разрывов (выходные, без День/План/Факт заголовка)
    в диапазоне календарной части листа.

    НАЙДЕНО 31.08.2026 (файл «...от 30.08.26.xlsx»): merge-диапазон
    заголовка месяца (row2) есть только у ОКТЯБРЯ (GZ2:NL2) — у ИЮНЯ,
    ИЮЛЯ, АВГУСТА, СЕНТЯБРЯ название месяца стоит текстом в одной
    незамерженной ячейке (проверено по сырому XML sheet1.xml — merge
    для этих месяцев в файле физически отсутствует, не артефакт починки
    styles.xml, которая этот лист не трогает). Прежняя версия строила
    month_ranges только по merged_cells.ranges — из-за этого 4 месяца
    из 5 молча выпадали из календаря (парсились только 104 колонки
    ОКТЯБРЯ вместо всех ~287). Теперь границы месяца строятся по самим
    позициям текста в row2 (работает одинаково для мерженых и
    немерженых ячеек — сам текст присутствует в обоих случаях), конец
    календаря — по последней колонке с непустым row3/row4, а не по
    границе merge-диапазона (у последнего месяца её и не было).
    """
    max_col = max_col or ws.max_column
    month_starts = []  # (start_col, month_num), по факту наличия текста в row2
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip().upper() in MONTHS_RU:
            month_starts.append((c, MONTHS_RU[v.strip().upper()]))
    if not month_starts:
        return [], (None, None)
    month_starts.sort()
    calendar_start = month_starts[0][0]
    calendar_end = calendar_start
    for c in range(calendar_start, max_col + 1):
        if ws.cell(row=day_row, column=c).value is not None or norm_header(ws.cell(row=subhdr_row, column=c).value):
            calendar_end = c

    def month_for_col(c):
        m = None
        for start, mm in month_starts:
            if start <= c:
                m = mm
            else:
                break
        return m

    columns = []
    gap_columns = []
    col = calendar_start
    while col <= calendar_end:
        day_val = ws.cell(row=day_row, column=col).value
        plan_hdr = norm_header(ws.cell(row=subhdr_row, column=col).value)
        # НАЙДЕНО 29.08.2026 (блокер к запуску 01.09): наличие числа в
        # строке "день" (row3) само по себе не значит, что колонка —
        # плановая. Воскресные "безголовые" колонки (выходной, значение
        # "В") ТОЖЕ несут день-номер в строке 3 (например, "5" для
        # 05.07.2026), но в строке 4 (subhdr_row) у них нет "План" —
        # там просто пусто. Без проверки заголовка строки 4 эти колонки
        # ошибочно принимались за плановые, и "В" улетало в
        # daily_progress как текстовое значение planned_crew. Теперь
        # колонка считается плановой, только если И день есть, И
        # заголовок строки 4 — именно "план".
        if isinstance(day_val, (int, float)) and "план" in plan_hdr:
            month = month_for_col(col)
            day = int(day_val)
            columns.append({"col": col, "kind": "plan", "year": DEFAULT_YEAR, "month": month, "day": day})
            fact_col = col + 1
            fact_hdr = norm_header(ws.cell(row=subhdr_row, column=fact_col).value)
            if "факт" in fact_hdr:
                columns.append({"col": fact_col, "kind": "fact", "year": DEFAULT_YEAR, "month": month, "day": day})
                col = fact_col + 1
            else:
                col += 1
        else:
            gap_columns.append(col)
            col += 1
    return columns, gap_columns


def iter_data_rows(ws, start_row, meta_col_max):
    """Строки данных, начиная с start_row, до первого 3-кратного пустого разрыва в колонке A."""
    empty_streak = 0
    r = start_row
    while r <= ws.max_row:
        a_val = ws.cell(row=r, column=1).value
        row_has_any = any(
            ws.cell(row=r, column=c).value is not None for c in range(1, meta_col_max + 1)
        )
        if a_val is None and not row_has_any:
            empty_streak += 1
            if empty_streak >= 3:
                break
        else:
            empty_streak = 0
            yield r
        r += 1


def is_section_header_row(ws, row, name_col):
    a = ws.cell(row=row, column=1).value
    name = ws.cell(row=row, column=name_col).value if name_col else None
    if a is None:
        return False
    if isinstance(a, str) and not a.strip().isdigit():
        # текст в колонке A, остальное метаполе (имя работы) пусто
        if name is None:
            return True
    return False


def is_totals_row(ws, row):
    a = ws.cell(row=row, column=1).value
    return isinstance(a, str) and "итого" in a.lower()


def normalize_numeric_or_flag(raw):
    if raw is None:
        return None, None
    if isinstance(raw, (int, float)):
        return float(raw), str(raw)
    s = str(raw).strip()
    if s == "":
        return None, None
    m = re.match(r"^(\d+(?:[.,]\d+)?)\s*шт\.?$", s, re.IGNORECASE)
    if m:
        # физический объём в штуках, не процент
        return None, s
    m = re.match(r"^(\d+(?:[.,]\d+)?)$", s)
    if m:
        return float(m.group(1).replace(",", ".")), s
    return None, s


def parse_sheet(ws, sheet_name, cfg, code_prefix, quality_issues):
    fields, missing = find_header_columns(ws, cfg["fields"])
    if "name" not in fields:
        raise RuntimeError(f"[{sheet_name}] не найдена колонка «Наименование работ» — проверить шапку (см. skill)")

    calendar_cols, gap_cols = build_calendar_columns(ws)
    meta_col_max = max(fields.values(), default=9) + 1

    works = []
    daily = []
    seq = 0
    current_section = None
    started = False

    name_col = fields["name"]

    for row in iter_data_rows(ws, start_row=5, meta_col_max=meta_col_max):
        if is_totals_row(ws, row):
            # итоговая строка "Итого кол-во человек" — не работа, обрабатывается отдельно
            continue
        if is_section_header_row(ws, row, name_col):
            current_section = str(ws.cell(row=row, column=1).value).strip()
            started = True
            continue
        name_val = ws.cell(row=row, column=name_col).value
        if name_val is None or (isinstance(name_val, str) and not name_val.strip()):
            # "Наименование работ" пусто — на листе РСК встречается отдельно от
            # "Содержание работы" (описание дефекта), которое может быть заполнено,
            # когда наименование не указано. Не терять строку молча.
            content_col = fields.get("content")
            fallback = ws.cell(row=row, column=content_col).value if content_col else None
            if fallback is None or (isinstance(fallback, str) and not fallback.strip()):
                continue
            name_val = fallback
            row_quality_notes = [
                f"«Наименование работ» пусто ({get_column_letter(name_col)}{row}), "
                f"использовано «Содержание работы» — сверить с ПТО ИКС"
            ]
            quality_issues.append({
                "sheet": sheet_name, "row": row, "type": "name_missing_used_content_fallback",
                "cell": f"{get_column_letter(name_col)}{row}",
            })
        else:
            row_quality_notes = []

        seq += 1
        work_source = cfg["source"]
        if current_section and AUX_SECTION_TITLE in current_section.lower():
            work_source = "aux"

        code = f"TM35-{code_prefix[work_source]}-{seq:03d}"

        crew_raw = ws.cell(row=row, column=fields["crew"]).value if "crew" in fields else None
        if isinstance(crew_raw, date_cls):
            # НАЙДЕНО 31.08.2026 (строка 108, «Электрофикация УТ13»): Excel
            # автозаменил похожее на "5.4"/"4-5" значение "Кол-во чел." на
            # дату — настоящее число людей отсюда не восстановить, не
            # угадываем. Сохраняем сырую дату как текст для ручной сверки
            # с ПТО, не даём datetime уйти в JSON как есть (не сериализуется).
            crew_dt = crew_raw
            crew_raw = f"похоже на дату: {crew_dt.strftime('%d.%m.%Y')}"
            quality_issues.append({
                "sheet": sheet_name, "row": row, "type": "crew_looks_like_date",
                "value": crew_dt.isoformat(), "cell": f"{get_column_letter(fields['crew'])}{row}",
            })
            row_quality_notes.append(
                f"«Кол-во чел.» = дата {crew_dt.strftime('%d.%m.%Y')} ({get_column_letter(fields['crew'])}{row}) — "
                f"похоже на «5.4»/«4-5», автоматически распознанное Excel как дата; "
                f"реальное число не восстановлено, сверить с ПТО"
            )
        executor_type = "own_forces"
        subcontractor_note = None
        if isinstance(crew_raw, str) and "подряд" in crew_raw.lower():
            executor_type = "subcontract"
            note_col = fields.get("note") or fields.get("equipment")
            subcontractor_note = ws.cell(row=row, column=note_col).value if note_col else None

        pct_val = ws.cell(row=row, column=fields["pct"]).value if "pct" in fields else None
        pct_norm, pct_raw = normalize_numeric_or_flag(pct_val)
        if pct_val is not None and pct_norm is None and pct_raw:
            quality_issues.append({
                "sheet": sheet_name, "row": row, "type": "pct_not_numeric",
                "value": pct_raw, "cell": f"{get_column_letter(fields['pct'])}{row}",
                "work_code": code,
            })
            row_quality_notes.append(
                f"«% вып.» = {pct_raw!r} не число ({get_column_letter(fields['pct'])}{row}) — "
                f"похоже на физический объём в штуках, не процент; не нормализовано автоматически"
            )

        volume_val = ws.cell(row=row, column=fields["volume"]).value if "volume" in fields else None
        volume_norm, volume_raw = normalize_numeric_or_flag(volume_val)
        if volume_val is not None and volume_norm is None and volume_raw:
            quality_issues.append({
                "sheet": sheet_name, "row": row, "type": "volume_not_numeric",
                "value": volume_raw, "cell": f"{get_column_letter(fields['volume'])}{row}",
                "work_code": code,
            })
            row_quality_notes.append(
                f"«Объём работ» = {volume_raw!r} не число ({get_column_letter(fields['volume'])}{row}) — "
                f"не нормализовано автоматически"
            )

        name_str = clean_text_field(str(name_val)) or str(name_val).strip()
        work_rec = {
            "code": code,
            "source": work_source,
            "section": current_section,
            "name": name_str,
            "location": guess_location(name_str),
            "unit": clean_text_field(ws.cell(row=row, column=fields["unit"]).value if "unit" in fields else None),
            "volume": volume_norm,
            "volume_raw": volume_raw,
            "crew_raw": crew_raw,
            "executor_type": executor_type,
            "status": derive_status(pct_norm),
            "baseline_finish": parse_baseline_month(
                ws.cell(row=row, column=fields["month_hint"]).value if "month_hint" in fields else None
            ),
            "baseline_crew": parse_crew_number(crew_raw),
            "equipment": clean_text_field(ws.cell(row=row, column=fields["equipment"]).value if "equipment" in fields else None),
            "comment": clean_text_field(ws.cell(row=row, column=fields["note"]).value if "note" in fields else None),
            "subcontractor_note": clean_text_field(subcontractor_note),
            "fact_pct": pct_norm,
            "fact_pct_raw": pct_raw,
            "source_row_ref": f"{sheet_name}!{row}",
            "data_quality_flag": "needs_review" if row_quality_notes else "ok",
            "data_quality_note": "; ".join(row_quality_notes) or None,
        }
        works.append(work_rec)

        # календарные значения + комментарии по этой строке
        for cc in calendar_cols:
            cell = ws.cell(row=row, column=cc["col"])
            comment_text = clean_comment_text(cell.comment.text) if cell.comment else None
            cell_value = cell.value
            if isinstance(cell_value, str) and not cell_value.strip():
                # НАЙДЕНО 31.08.2026: массово (17378 из 18319 в этом файле)
                # ячейка визуально пустая, но cell.value == '\xa0' (одинокий
                # неразрывный пробел, не None) — тот же симптом, что баг
                # парсера 14.08 (excel_import, все поля плана/факта пустые),
                # другая причина. .strip() убирает и \xa0 (Unicode-пробел),
                # так что визуально-пустая ячейка теперь и технически пустая
                # — строка в daily_progress не создаётся, если нет и
                # комментария. Не угадывание значения — просто не считаем
                # пробел значением.
                cell_value = None
            if cell_value is None and comment_text is None:
                continue
            try:
                iso_date = f"{cc['year']:04d}-{cc['month']:02d}-{cc['day']:02d}"
            except (TypeError, ValueError):
                # Дата не определяется однозначно (месяц не резолвится для этой
                # колонки) — валидного `date` для NOT NULL колонки daily_progress.date
                # нет, вставлять строку с угаданной датой нельзя (РЕШЕНИЯ_v1.1:
                # "не угадывать"). Уходит в import_unresolved_cell, не в daily_progress.
                quality_issues.append({
                    "sheet": sheet_name, "row": row, "type": "bad_calendar_date",
                    "cell": f"{get_column_letter(cc['col'])}{row}", "raw": cc,
                    "work_code": code, "value": cell_value,
                })
                continue
            daily.append({
                "work_code": code,
                "date": iso_date,
                "kind": cc["kind"],   # plan | fact
                "value": cell_value,
                "comment": comment_text,
                "cell": f"{get_column_letter(cc['col'])}{row}",
                "data_quality_flag": "ok",
                "data_quality_note": None,
            })

    # день-уровневые комментарии на строке 3 (заголовок дня) — не привязаны к работе.
    # Проверяем обе колонки пары (План И Факт) — комментарий встречается на любой
    # из них, дедуп по (дата, текст) на случай, если стоит на обеих.
    day_level_comments = []
    seen_day_comments = set()
    for cc in calendar_cols:
        cell = ws.cell(row=3, column=cc["col"])
        if cell.comment:
            date_str = f"{cc['year']:04d}-{cc['month']:02d}-{cc['day']:02d}"
            text = clean_comment_text(cell.comment.text)
            if text is None:
                continue
            key = (date_str, text)
            if key in seen_day_comments:
                continue
            seen_day_comments.add(key)
            day_level_comments.append({
                "sheet": sheet_name,
                "date": date_str,
                "comment": text,
                "blocker_type": guess_blocker_type(text),
            })

    return works, daily, day_level_comments


def main(xlsx_path, out_dir):
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    code_prefix = {"main": "MAIN", "aux": "AUX"}

    all_works, all_daily, all_day_comments = [], [], []
    quality_issues = []

    for sheet_name, cfg in SHEET_CONFIGS.items():
        if sheet_name not in wb.sheetnames:
            quality_issues.append({"sheet": sheet_name, "type": "sheet_missing"})
            continue
        ws = wb[sheet_name]
        works, daily, day_comments = parse_sheet(ws, sheet_name, cfg, code_prefix, quality_issues)
        all_works.extend(works)
        all_daily.extend(daily)
        all_day_comments.extend(day_comments)

    # Matrix-baseline (docs/MATRIX_DIAGNOSTICS.md) — точнее текстового
    # "Месяц выполнения работ", подменяет его там, где есть календарные
    # данные; иначе явный source='text_month_only'/'no_data', не молчим.
    matrix_baselines = compute_matrix_baselines(all_daily)
    for w in all_works:
        mb = matrix_baselines.get(w["code"])
        if mb:
            w["baseline_start"] = mb["plan_start"]
            w["baseline_finish"] = mb["plan_finish"]  # переопределяет текстовый месяц точной датой
            w["baseline_source"] = mb["source"]
            w["baseline_confidence"] = mb["confidence"]
        elif w.get("baseline_finish"):
            w["baseline_start"] = None
            w["baseline_source"] = "text_month_only"
            w["baseline_confidence"] = "low"
        else:
            w["baseline_start"] = None
            w["baseline_source"] = "no_data"
            w["baseline_confidence"] = "none"

    (out_dir / "works.json").write_text(
        json.dumps(all_works, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out_dir / "daily_progress.json").write_text(
        json.dumps(all_daily, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out_dir / "day_level_comments.json").write_text(
        json.dumps(all_day_comments, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    (out_dir / "quality_report.json").write_text(
        json.dumps(quality_issues, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    by_source = defaultdict(int)
    for w in all_works:
        by_source[w["source"]] += 1

    print(f"Работ разобрано: {len(all_works)}  {dict(by_source)}")
    print(f"Записей план/факт: {len(all_daily)}")
    print(f"День-уровневых комментариев: {len(all_day_comments)}")
    print(f"Замечаний качества данных: {len(quality_issues)}")
    print(f"Результат: {out_dir}/")


if __name__ == "__main__":
    xlsx = sys.argv[1] if len(sys.argv) > 1 else \
        "/home/oleg/Documents/TM-35/import/smr_300826/Сводная Таблица по ТМ 35 от 30.08.26_fixed.xlsx"
    out = sys.argv[2] if len(sys.argv) > 2 else "/home/oleg/Documents/TM-35/import/smr_300826/output"
    main(xlsx, out)
